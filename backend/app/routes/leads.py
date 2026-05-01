"""
Lead Routes - CRUD operations for leads and lead management.
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.models import Lead, LeadStatus, LeadSource, LeadActivity
from app.models.schemas import LeadCreate, LeadUpdate, LeadResponse
from app.services.websocket_manager import ws_manager

router = APIRouter(prefix="/api/leads", tags=["leads"])


@router.get("", response_model=list[LeadResponse])
async def get_leads(
    status: Optional[str] = None,
    source: Optional[str] = None,
    min_score: Optional[float] = None,
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Lead)

    if status:
        stmt = stmt.where(Lead.status == status)
    if source:
        stmt = stmt.where(Lead.source == source)
    if min_score is not None:
        stmt = stmt.where(Lead.score >= min_score)
    if search:
        search_filter = f"%{search}%"
        stmt = stmt.where(
            (Lead.first_name.ilike(search_filter))
            | (Lead.last_name.ilike(search_filter))
            | (Lead.email.ilike(search_filter))
            | (Lead.company.ilike(search_filter))
        )

    # Sorting
    sort_col = getattr(Lead, sort_by, Lead.created_at)
    if sort_order == "asc":
        stmt = stmt.order_by(sort_col.asc())
    else:
        stmt = stmt.order_by(sort_col.desc())

    stmt = stmt.offset(skip).limit(limit)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.get("/count")
async def get_leads_count(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(func.count(Lead.id)))
    total = result.scalar()

    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    result_today = await db.execute(
        select(func.count(Lead.id)).where(Lead.created_at >= today)
    )
    today_count = result_today.scalar()

    result_qualified = await db.execute(
        select(func.count(Lead.id)).where(Lead.status == LeadStatus.QUALIFIED)
    )
    qualified = result_qualified.scalar()

    return {
        "total": total,
        "today": today_count,
        "qualified": qualified,
    }


@router.get("/stats")
async def get_lead_stats(db: AsyncSession = Depends(get_db)):
    """Get comprehensive lead statistics."""
    # By status
    status_result = await db.execute(
        select(Lead.status, func.count(Lead.id)).group_by(Lead.status)
    )
    by_status = {
        (row[0].value if hasattr(row[0], "value") else str(row[0])): row[1]
        for row in status_result.all()
    }

    # By source
    source_result = await db.execute(
        select(Lead.source, func.count(Lead.id)).group_by(Lead.source)
    )
    by_source = {
        (row[0].value if hasattr(row[0], "value") else str(row[0])): row[1]
        for row in source_result.all()
    }

    # Leads over time (last 30 days)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    daily_result = await db.execute(
        select(
            func.date(Lead.created_at),
            func.count(Lead.id),
        )
        .where(Lead.created_at >= thirty_days_ago)
        .group_by(func.date(Lead.created_at))
        .order_by(func.date(Lead.created_at))
    )
    daily_leads = [{"date": str(row[0]), "count": row[1]} for row in daily_result.all()]

    # Average score
    avg_result = await db.execute(select(func.avg(Lead.score)))
    avg_score = round(avg_result.scalar() or 0, 2)

    # Conversion rate
    total_result = await db.execute(select(func.count(Lead.id)))
    total = total_result.scalar() or 1
    won_result = await db.execute(
        select(func.count(Lead.id)).where(Lead.status == LeadStatus.CLOSED_WON)
    )
    won = won_result.scalar() or 0
    conversion_rate = round((won / total) * 100, 1) if total > 0 else 0

    return {
        "by_status": by_status,
        "by_source": by_source,
        "daily_leads": daily_leads,
        "average_score": avg_score,
        "conversion_rate": conversion_rate,
        "total": total,
    }


@router.post("/enrich")
async def enrich_leads_endpoint(
    batch_size: int = Query(20, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
):
    """Enrich leads with LinkedIn URLs, phone numbers, and emails."""
    from app.services.lead_enrichment import enrich_leads
    result = await enrich_leads(db, batch_size=batch_size)
    return result


@router.get("/export/excel")
async def export_leads_excel(
    status: Optional[str] = None,
    min_score: Optional[float] = None,
    db: AsyncSession = Depends(get_db),
):
    """Export all leads as an Excel (.xlsx) report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    query = select(Lead).order_by(Lead.score.desc())
    if status:
        query = query.where(Lead.status == status)
    if min_score is not None:
        query = query.where(Lead.score >= min_score)

    result = await db.execute(query)
    leads = result.scalars().all()

    wb = Workbook()

    # --- Sheet 1: All Leads ---
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = [
        "Company", "First Name", "Last Name", "Email", "Phone",
        "LinkedIn", "Website", "Score (%)", "Status", "Service Needed",
        "Source", "Source URL", "Industry", "Location", "Company Size",
        "AI Summary", "AI Next Action", "Notes", "Tags",
        "Created At", "Last Contacted",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = "A2"

    score_green = Font(color="2E7D32", bold=True)
    score_yellow = Font(color="F57F17", bold=True)
    score_red = Font(color="C62828")

    for row_idx, lead in enumerate(leads, 2):
        service = (lead.custom_fields or {}).get("service_needed", "")
        score_pct = round(lead.score * 100)
        tags_str = ", ".join(lead.tags) if lead.tags else ""
        created = lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else ""
        contacted = lead.last_contacted_at.strftime("%Y-%m-%d %H:%M") if lead.last_contacted_at else ""

        row_data = [
            lead.company, lead.first_name, lead.last_name, lead.email, lead.phone,
            lead.linkedin_url, lead.website, score_pct,
            lead.status.value if hasattr(lead.status, 'value') else str(lead.status),
            service,
            lead.source.value if hasattr(lead.source, 'value') else str(lead.source),
            lead.source_url, lead.industry, lead.location, lead.company_size,
            lead.ai_summary, lead.ai_next_action, lead.notes, tags_str,
            created, contacted,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Color-code score
        score_cell = ws.cell(row=row_idx, column=8)
        if score_pct >= 40:
            score_cell.font = score_green
        elif score_pct >= 30:
            score_cell.font = score_yellow
        else:
            score_cell.font = score_red

        # Zebra striping
        if row_idx % 2 == 0:
            stripe_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = stripe_fill

    # Auto-width columns (capped)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=min(len(leads) + 1, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads) + 1}"

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = "1F4E79"

    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws2.cell(row=1, column=1, value="Open Claw Hermes Marketing - Lead Report").font = title_font
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(italic=True, color="666666")

    summary_header = Font(bold=True, size=11)
    ws2.cell(row=4, column=1, value="Metric").font = summary_header
    ws2.cell(row=4, column=2, value="Value").font = summary_header
    for c in [1, 2]:
        ws2.cell(row=4, column=c).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    total = len(leads)
    by_status = {}
    by_service = {}
    total_score = 0
    with_email = 0
    with_phone = 0
    with_linkedin = 0

    for lead in leads:
        by_status[lead.status] = by_status.get(lead.status, 0) + 1
        svc = (lead.custom_fields or {}).get("service_needed", "unknown")
        by_service[svc] = by_service.get(svc, 0) + 1
        total_score += lead.score
        if lead.email:
            with_email += 1
        if lead.phone:
            with_phone += 1
        if lead.linkedin_url:
            with_linkedin += 1

    avg_score = round((total_score / total) * 100) if total else 0

    summary_rows = [
        ("Total Leads", total),
        ("Average Score", f"{avg_score}%"),
        ("With Email", f"{with_email} ({round(with_email / max(total, 1) * 100)}%)"),
        ("With Phone", f"{with_phone} ({round(with_phone / max(total, 1) * 100)}%)"),
        ("With LinkedIn", f"{with_linkedin} ({round(with_linkedin / max(total, 1) * 100)}%)"),
        ("", ""),
        ("Status Breakdown", ""),
    ]
    for s, c in sorted(by_status.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {s.capitalize()}", c))

    summary_rows.append(("", ""))
    summary_rows.append(("Service Breakdown", ""))
    for s, c in sorted(by_service.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {s.replace('_', ' ').title()}", c))

    for i, (metric, value) in enumerate(summary_rows, 5):
        ws2.cell(row=i, column=1, value=metric)
        ws2.cell(row=i, column=2, value=value)
        if metric and not metric.startswith("  ") and metric != "":
            ws2.cell(row=i, column=1).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"hermes_marketing_leads_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{lead_id}", response_model=LeadResponse)
async def get_lead(lead_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return lead


@router.post("", response_model=LeadResponse)
async def create_lead(lead_data: LeadCreate, db: AsyncSession = Depends(get_db)):
    lead = Lead(**lead_data.model_dump())
    db.add(lead)
    await db.commit()
    await db.refresh(lead)

    await ws_manager.broadcast({
        "event": "new_lead",
        "data": {"lead_id": lead.id, "email": lead.email, "company": lead.company},
    })

    return lead


@router.put("/{lead_id}", response_model=LeadResponse)
async def update_lead(lead_id: str, update_data: LeadUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    for field, value in update_data.model_dump(exclude_unset=True).items():
        setattr(lead, field, value)
    lead.updated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(lead)
    return lead


@router.delete("/{lead_id}")
async def delete_lead(lead_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    await db.delete(lead)
    await db.commit()
    return {"success": True, "message": "Lead deleted"}


@router.get("/{lead_id}/activities")
async def get_lead_activities(lead_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(LeadActivity)
        .where(LeadActivity.lead_id == lead_id)
        .order_by(LeadActivity.created_at.desc())
    )
    activities = result.scalars().all()
    return [
        {
            "id": a.id,
            "activity_type": a.activity_type,
            "description": a.description,
            "extra_data": a.extra_data,
            "created_at": a.created_at.isoformat(),
        }
        for a in activities
    ]
